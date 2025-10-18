from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import time
import os
import re
import sys
import traceback

class CarPriceTracker:
    def __init__(self, gmail_user, gmail_app_password, recipient_email, max_retries=3):
        """
        GitHub Actions için optimize edilmiş tracker

        Args:
            gmail_user: Gmail adresiniz
            gmail_app_password: Gmail App Password
            recipient_email: Raporun gönderileceği email
            max_retries: Hata durumunda maksimum deneme sayısı
        """
        self.gmail_user = gmail_user
        self.gmail_app_password = gmail_app_password
        self.recipient_email = recipient_email
        self.max_retries = max_retries
        self.excel_filename = f'arac_fiyatlari_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        self.success_count = 0
        self.fail_count = 0

    def setup_driver(self):
        """GitHub Actions için Chrome WebDriver yapılandırması"""
        chrome_options = Options()

        # GitHub Actions için gerekli ayarlar
        chrome_options.add_argument('--headless=new')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument('--disable-popup-blocking')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_experimental_option("prefs", {
            "profile.default_content_setting_values.notifications": 2
        })

        # Loglama seviyesini azalt
        chrome_options.add_argument('--log-level=3')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.set_page_load_timeout(30)
            return driver
        except Exception as e:
            print(f"❌ ChromeDriver hatası: {e}")
            print("Alternatif yol deneniyor...")
            service = Service('/usr/local/bin/chromedriver')
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(30)
            return driver

    def clean_price(self, price_text):
        """Fiyat metnini temizler ve formatlar"""
        if not price_text or price_text == "Fiyat sitede bulunamadı":
            return None

        # Sadece rakamları al
        numbers = re.sub(r'[^\d]', '', price_text)

        if numbers:
            # Binlik ayraç ekle
            try:
                formatted = "{:,}".format(int(numbers)).replace(',', '.')
                return formatted + ' TL'
            except:
                return numbers

        return price_text

    def close_popups(self, driver):
        """Çerez ve diğer popup'ları kapat"""
        popup_selectors = [
            "//button[contains(text(), 'Kabul')]",
            "//button[contains(text(), 'Accept')]",
            "//button[contains(text(), 'Tümünü kabul')]",
            "//button[contains(@class, 'cookie')]",
            "//button[contains(@id, 'accept')]",
            "//a[contains(@class, 'close')]",
            "//button[contains(@class, 'close')]"
        ]

        for selector in popup_selectors:
            try:
                element = driver.find_element(By.XPATH, selector)
                element.click()
                time.sleep(0.5)
            except:
                pass

    def extract_price(self, driver):
        """Sayfadan fiyat bilgisini çıkar - gelişmiş selector'lar"""
        price_selectors = [
            # Fiyat div/span'leri
            "//div[contains(@class, 'price') and contains(text(), 'TL')]",
            "//span[contains(@class, 'price') and contains(text(), 'TL')]",
            "//p[contains(@class, 'price') and contains(text(), 'TL')]",

            # TL içeren elementler
            "//*[contains(text(), 'TL') and not(contains(text(), 'İletişim'))]",
            "//*[contains(text(), '₺')]",

            # Binlik ayraç olanlar
            "//*[contains(text(), '.000')]",
            "//*[contains(text(), '.500')]",
            "//*[contains(text(), '.999')]",

            # Model price class'ları
            "//div[contains(@class, 'model-price')]",
            "//div[contains(@class, 'vehicle-price')]",
            "//div[contains(@class, 'car-price')]",

            # Data attribute'lar
            "//*[@data-price]",
            "//*[@data-vehicle-price]",
        ]

        found_prices = []

        for selector in price_selectors:
            try:
                elements = driver.find_elements(By.XPATH, selector)
                for elem in elements:
                    text = elem.text.strip()

                    # Fiyat gibi görünen metin mi?
                    if text and any(c.isdigit() for c in text):
                        if 'TL' in text or '₺' in text or '.000' in text:
                            # İstenmeyen metinleri filtrele
                            if not any(x in text.lower() for x in ['ödeme', 'taksit', 'ay', 'kredi', 'iletişim']):
                                found_prices.append(text)

                if found_prices:
                    break

            except Exception as e:
                continue

        if found_prices:
            # En uzun ve en bilgilendirici olanı seç
            best_price = max(found_prices, key=lambda x: len(x) if len(x) < 100 else 0)
            return best_price

        return None

    def scrape_model_with_retry(self, model_info, brand):
        """Tek bir modeli retry mekanizması ile çek"""
        for attempt in range(self.max_retries):
            driver = None
            try:
                driver = self.setup_driver()

                print(f"   → {model_info['name']} (Deneme {attempt + 1}/{self.max_retries})")

                driver.get(model_info['url'])
                time.sleep(4)

                # Popup'ları kapat
                self.close_popups(driver)

                # Sayfanın tamamen yüklenmesini bekle
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )

                # Fiyat çıkar
                price_text = self.extract_price(driver)

                if price_text:
                    self.success_count += 1
                    print(f"      ✓ Başarılı: {price_text[:50]}")

                    return {
                        'Marka': brand,
                        'Model': model_info['name'],
                        'Fiyat': price_text,
                        'Fiyat (Temiz)': self.clean_price(price_text),
                        'Tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Durum': 'Başarılı',
                        'URL': model_info['url']
                    }
                else:
                    if attempt < self.max_retries - 1:
                        print(f"      ⚠️  Fiyat bulunamadı, tekrar deneniyor...")
                        time.sleep(2)
                        continue
                    else:
                        print(f"      ✗ Fiyat bulunamadı (Tüm denemeler tükendi)")
                        self.fail_count += 1
                        return {
                            'Marka': brand,
                            'Model': model_info['name'],
                            'Fiyat': 'Fiyat sitede bulunamadı',
                            'Fiyat (Temiz)': 'N/A',
                            'Tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'Durum': 'Başarısız - Fiyat Bulunamadı',
                            'URL': model_info['url']
                        }

            except TimeoutException:
                print(f"      ⏱️  Timeout - Sayfa yüklenemedi")
                if attempt < self.max_retries - 1:
                    time.sleep(3)
                    continue

            except Exception as e:
                print(f"      ✗ Hata: {str(e)[:50]}")
                if attempt < self.max_retries - 1:
                    time.sleep(2)
                    continue

            finally:
                if driver:
                    driver.quit()
                time.sleep(2)

        # Tüm denemeler başarısız
        self.fail_count += 1
        return {
            'Marka': brand,
            'Model': model_info['name'],
            'Fiyat': 'Hata: Tüm denemeler başarısız',
            'Fiyat (Temiz)': 'N/A',
            'Tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Durum': 'Başarısız - Hata',
            'URL': model_info['url']
        }

    def scrape_hyundai_prices(self):
        """Hyundai fiyatları - retry mekanizmalı"""
        print("🚗 Hyundai fiyatları çekiliyor...")
        print("="*60)

        models_info = [
            {'name': 'i10', 'url': 'https://www.hyundai.com/tr/tr/modeller/i10/satinal'},
            {'name': 'i20', 'url': 'https://www.hyundai.com/tr/tr/modeller/i20/satinal'},
            {'name': 'Bayon', 'url': 'https://www.hyundai.com/tr/tr/modeller/bayon/satinal'},
            {'name': 'Kona', 'url': 'https://www.hyundai.com/tr/tr/modeller/kona/satinal'},
            {'name': 'Kona Electric', 'url': 'https://www.hyundai.com/tr/tr/modeller/kona-electric/satinal'},
            {'name': 'Tucson', 'url': 'https://www.hyundai.com/tr/tr/modeller/tucson/satinal'},
            {'name': 'Santa Fe', 'url': 'https://www.hyundai.com/tr/tr/modeller/santa-fe/satinal'},
            {'name': 'Ioniq 5', 'url': 'https://www.hyundai.com/tr/tr/modeller/ioniq-5/satinal'},
            {'name': 'Ioniq 6', 'url': 'https://www.hyundai.com/tr/tr/modeller/ioniq-6/satinal'},
        ]

        hyundai_data = []
        for model in models_info:
            result = self.scrape_model_with_retry(model, 'Hyundai')
            hyundai_data.append(result)

        print("="*60)
        print(f"✅ Hyundai tamamlandı: {len(hyundai_data)} model işlendi\n")
        return hyundai_data

    def scrape_kia_prices(self):
        """Kia fiyatları - retry mekanizmalı"""
        print("🚗 Kia fiyatları çekiliyor...")
        print("="*60)

        models_info = [
            {'name': 'Picanto', 'url': 'https://www.kia.com/tr/modeller/yeni-picanto/satin-al.html'},
            {'name': 'Stonic', 'url': 'https://www.kia.com/tr/modeller/stonic/satin-al.html'},
            {'name': 'Ceed', 'url': 'https://www.kia.com/tr/modeller/ceed-hb/satin-al.html'},
            {'name': 'XCeed', 'url': 'https://www.kia.com/tr/modeller/xceed/satin-al.html'},
            {'name': 'Sportage', 'url': 'https://www.kia.com/tr/modeller/sportage-nq5/satin-al.html'},
            {'name': 'Sorento', 'url': 'https://www.kia.com/tr/modeller/sorento/satin-al.html'},
            {'name': 'EV3', 'url': 'https://www.kia.com/tr/modeller/ev3/satin-al.html'},
            {'name': 'EV6', 'url': 'https://www.kia.com/tr/modeller/ev6/satin-al.html'},
            {'name': 'EV9', 'url': 'https://www.kia.com/tr/modeller/ev9/satin-al.html'},
        ]

        kia_data = []
        for model in models_info:
            result = self.scrape_model_with_retry(model, 'Kia')
            kia_data.append(result)

        print("="*60)
        print(f"✅ Kia tamamlandı: {len(kia_data)} model işlendi\n")
        return kia_data

    def save_to_excel(self, data):
        """Excel'e kaydet - gelişmiş formatlama"""
        if not data:
            print("⚠️  Kaydedilecek veri yok!")
            return False

        try:
            df = pd.DataFrame(data)

            # Sütun sıralaması
            columns_order = ['Marka', 'Model', 'Fiyat', 'Fiyat (Temiz)', 'Durum', 'Tarih', 'URL']
            df = df[columns_order]

            with pd.ExcelWriter(self.excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fiyatlar', index=False)

                workbook = writer.book
                worksheet = writer.sheets['Fiyatlar']

                # Sütun genişlikleri
                column_widths = {
                    'A': 12,  # Marka
                    'B': 18,  # Model
                    'C': 35,  # Fiyat
                    'D': 20,  # Fiyat (Temiz)
                    'E': 25,  # Durum
                    'F': 20,  # Tarih
                    'G': 60,  # URL
                }

                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # Başlık formatı
                from openpyxl.styles import Font, PatternFill, Alignment

                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Durum sütunu renklendirme
                for row in range(2, len(df) + 2):
                    status_cell = worksheet[f'E{row}']
                    if 'Başarılı' in str(status_cell.value):
                        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif 'Başarısız' in str(status_cell.value):
                        status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            print(f"✅ Excel dosyası oluşturuldu: {self.excel_filename}")
            print(f"   📊 Başarılı: {self.success_count} | Başarısız: {self.fail_count}")
            return True

        except Exception as e:
            print(f"❌ Excel kayıt hatası: {e}")
            traceback.print_exc()
            return False

    def send_email(self):
        """Gmail ile gönder - gelişmiş hata yönetimi"""
        if not os.path.exists(self.excel_filename):
            print("❌ Excel dosyası bulunamadı!")
            return False

        try:
            msg = MIMEMultipart()
            msg['From'] = self.gmail_user
            msg['To'] = self.recipient_email
            msg['Subject'] = f'🚗 Hyundai & Kia Fiyat Raporu - {datetime.now().strftime("%d.%m.%Y %H:%M")}'

            # Başarı oranı hesapla
            total = self.success_count + self.fail_count
            success_rate = (self.success_count / total * 100) if total > 0 else 0

            body = f"""
Merhaba,

{datetime.now().strftime("%d.%m.%Y %H:%M")} tarihli Hyundai ve Kia sıfır araç fiyat raporu ektedir.

📊 Özet:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ Başarılı: {self.success_count} model
❌ Başarısız: {self.fail_count} model
📈 Başarı Oranı: %{success_rate:.1f}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📦 Rapor İçeriği:
• Hyundai modelleri (9 model)
• Kia modelleri (9 model)
• Güncel fiyatlar ve linkler
• Durum bilgileri

Bu rapor GitHub Actions ile otomatik oluşturulmuştur.

İyi günler! 🚗

---
Otomatik Araç Fiyat Takip Sistemi
GitHub Actions - Powered by Selenium
            """

            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            # Excel dosyasını ekle
            with open(self.excel_filename, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={self.excel_filename}')
                msg.attach(part)

            # Gmail SMTP
            print("📧 Email gönderiliyor...")
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.set_debuglevel(0)
            server.starttls()
            server.login(self.gmail_user, self.gmail_app_password)
            server.send_message(msg)
            server.quit()

            print(f"✅ Email başarıyla gönderildi: {self.recipient_email}")
            return True

        except smtplib.SMTPAuthenticationError:
            print("❌ Gmail giriş hatası! App Password'ü kontrol edin.")
            return False
        except Exception as e:
            print(f"❌ Email gönderme hatası: {str(e)}")
            traceback.print_exc()
            return False

    def run(self):
        """Ana çalıştırma fonksiyonu"""
        print("\n" + "="*60)
        print("🚀 GITHUB ACTIONS - ARAÇ FİYAT TAKİP SİSTEMİ")
        print("="*60)
        print(f"⏰ Başlangıç: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*60 + "\n")

        start_time = time.time()

        try:
            # Fiyatları çek
            print("📥 Veri çekme işlemi başlıyor...\n")
            hyundai_data = self.scrape_hyundai_prices()
            kia_data = self.scrape_kia_prices()

            # Birleştir
            all_data = hyundai_data + kia_data

            if not all_data:
                print("❌ Hiç veri çekilemedi!")
                sys.exit(1)

            # Excel'e kaydet
            print("💾 Excel raporu oluşturuluyor...")
            if not self.save_to_excel(all_data):
                print("❌ Excel oluşturulamadı!")
                sys.exit(1)

            # Email gönder
            print("\n📧 Email gönderiliyor...")
            if not self.send_email():
                print("⚠️  Email gönderilemedi ama Excel oluşturuldu!")

            elapsed = time.time() - start_time

            print("\n" + "="*60)
            print("✅ İŞLEM TAMAMLANDI!")
            print("="*60)
            print(f"⏱️  Toplam Süre: {elapsed:.1f} saniye")
            print(f"✅ Başarılı: {self.success_count}")
            print(f"❌ Başarısız: {self.fail_count}")
            print(f"📊 Başarı Oranı: %{(self.success_count/(self.success_count+self.fail_count)*100):.1f}")
            print("="*60 + "\n")

            # Exit code belirle
            if self.fail_count > len(all_data) / 2:
                print("⚠️  Çok fazla hata! Exit code: 1")
                sys.exit(1)

        except KeyboardInterrupt:
            print("\n⚠️  İşlem kullanıcı tarafından durduruldu!")
            sys.exit(1)

        except Exception as e:
            print(f"\n❌ KRITIK HATA: {str(e)}")
            traceback.print_exc()
            sys.exit(1)


if __name__ == "__main__":

    print("🔧 GitHub Actions Car Price Tracker v2.0")
    print("="*60 + "\n")

    # GitHub Secrets'tan environment variables al
    GMAIL_USER = os.getenv('GMAIL_USER')
    GMAIL_APP_PASSWORD = os.getenv('GMAIL_APP_PASSWORD')
    RECIPIENT_EMAIL = os.getenv('RECIPIENT_EMAIL')

    # Kontrol et
    if not all([GMAIL_USER, GMAIL_APP_PASSWORD, RECIPIENT_EMAIL]):
        print("❌ HATA: GitHub Secrets ayarlanmamış!")
        print("\n📝 Yapılması gerekenler:")
        print("   1. Repository Settings > Secrets and variables > Actions")
        print("   2. Şu secret'ları ekleyin:")
        print("      • GMAIL_USER (Gmail adresiniz)")
        print("      • GMAIL_APP_PASSWORD (16 haneli App Password)")
        print("      • RECIPIENT_EMAIL (Hedef email)")
        print("\n🔗 Gmail App Password: https://myaccount.google.com/apppasswords")
        sys.exit(1)

    print("✅ Environment variables yüklendi")
    print(f"   Gmail: {GMAIL_USER}")
    print(f"   Alıcı: {RECIPIENT_EMAIL}\n")

    # Tracker'ı başlat ve çalıştır
    tracker = CarPriceTracker(
        gmail_user=GMAIL_USER,
        gmail_app_password=GMAIL_APP_PASSWORD,
        recipient_email=RECIPIENT_EMAIL,
        max_retries=3
    )

    tracker.run()